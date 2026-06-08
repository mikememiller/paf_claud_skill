"""
================================================================================
 Syntax Corporation © 2026 — PAF Agent Factory — templates/deliverables/build_oci_bom_xlsx.py
 Version : 2.0.0   Build : 2026.06.08
--------------------------------------------------------------------------------
 SPEC-DRIVEN OCI Bill of Materials + consumption estimator (.xlsx). Reads
 build/spec.json (spec.bom[], spec.estimator) → branded workbook. Requires openpyxl.
   python build_oci_bom_xlsx.py --spec build/spec.json --out build --prefix Sample_
================================================================================
"""
from __future__ import annotations
import argparse, json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

NAVY = "0632A0"; MIST = "F4F7FC"; WHITE = "FFFFFF"; INK = "16203A"
hdr = Font(name="Calibri", bold=True, color=WHITE, size=12)
ttl = Font(name="Georgia", bold=True, color=NAVY, size=16)
bold = Font(name="Calibri", bold=True, color=INK)
base = Font(name="Calibri", color=INK)
fill_hdr = PatternFill("solid", fgColor=NAVY)
fill_mist = PatternFill("solid", fgColor=MIST)
thin = Side(style="thin", color="DCE4F4")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr; cell.fill = fill_hdr; cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")


def main(argv=None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--spec", default="build/spec.json")
    ap.add_argument("--out", default="build")
    ap.add_argument("--prefix", default="")
    a = ap.parse_args(argv)

    spec = json.loads(Path(a.spec).read_text())
    meta = spec["meta"]; foot = (spec.get("branding") or {}).get("footer", "Syntax Corporation © 2026 · Confidential")
    wb = Workbook()

    # Sheet 1: BOM
    ws = wb.active; ws.title = "OCI BOM"
    ws["A1"] = f"{meta['agent_name']} — OCI Bill of Materials"; ws["A1"].font = ttl
    ws["A2"] = foot; ws["A2"].font = base
    ws.append([]); ws.append(["Component", "Pricing model", "Notes"]); style_header(ws, 4, 3)
    for r in spec["bom"]:
        ws.append([r["component"], r["pricing"], r["notes"]])
    for row in ws.iter_rows(min_row=5, max_row=4 + len(spec["bom"]), max_col=3):
        for c in row: c.font = base; c.border = border; c.fill = fill_mist
    ws.column_dimensions["A"].width = 42; ws.column_dimensions["B"].width = 38; ws.column_dimensions["C"].width = 34

    # Sheet 2: Estimator
    est = spec["estimator"]; tiers = est["tiers"]
    es = wb.create_sheet("Estimator")
    es["A1"] = "OCI Run Cost + Managed Services — by volume tier"; es["A1"].font = ttl
    es.append([]); es.append(["Tier", "Txns/yr", "OCI/yr (Lic. Incl.)", "OCI/yr (BYOL)",
                              "Mgd Svc 12-mo/mo", "24-mo/mo", "36-mo/mo"]); style_header(es, 3, 7)
    for t in tiers:
        es.append([t["name"], t["volume"], t["oci_li"], t["oci_byol"], t["ms12"], t["ms24"], t["ms36"]])
    for row in es.iter_rows(min_row=4, max_row=3 + len(tiers), max_col=7):
        for c in row:
            c.font = base; c.border = border; c.fill = fill_mist
            if c.column >= 3:
                c.number_format = '"$"#,##0'; c.alignment = Alignment(horizontal="right")
            elif c.column == 2:
                c.number_format = '#,##0'; c.alignment = Alignment(horizontal="right")
    for col, w in zip("ABCDEFG", (24, 12, 18, 16, 17, 14, 14)):
        es.column_dimensions[col].width = w
    es.append([]); es.append(["Notes:", est.get("note", "")])
    es.cell(row=es.max_row, column=1).font = bold

    out = Path(a.out); out.mkdir(parents=True, exist_ok=True)
    path = out / f"{a.prefix}OCI_BOM.xlsx"
    wb.save(path); print("wrote", path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
